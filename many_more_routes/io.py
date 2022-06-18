from ctypes import alignment
import openpyxl
from openpyxl.styles import Alignment

from pathlib import Path

from typing import List
from typing import Dict
from typing import Optional
from typing import Union

from . models import OutputRecord


ALIGNMENT_ROTATE = Alignment(
    horizontal='left',
    vertical='bottom',
    text_rotation=55,
    shrink_to_fit=False,
    wrap_text=False,
    indent=0
)

ALIGNMENT_FLAT = Alignment(
    horizontal='general',
    text_rotation=0,
    shrink_to_fit=False,
    wrap_text=False,
    indent=0
)


def load_excel(file_path: Union[str, Path], sheet_name: Optional[str] = None) -> List[Dict]:
    """ Loads an excel to a list of dictionaries where the first row is the column headers """

    if isinstance(file_path, str):
        file_path = Path(file_path)
    
    if not file_path.exists():
        raise ValueError(f'File "{file_path}" does not exist')

    workbook = openpyxl.load_workbook(str(file_path.absolute()))

    if sheet_name:
        sheet = workbook[sheet_name]

    else:
        sheet = workbook.active

    headers = [col.value for col in sheet[1]]

    return [
        {
            header: col.value for (header, col) in zip(headers, sheet[row])
        } for row in range(2, sheet.max_row + 1)
    ]


def write_excel(records: List[OutputRecord], file_path: Union[str, Path]) -> None:
    if isinstance(file_path, str):
        file_path = str(Path(file_path).absolute())
    else:
        file_path = str(file_path.absolute())

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = records[0]._api

    counts: Dict = {}
    for record in records:
        if record._api in workbook.sheetnames and record._api in counts.keys():
            sheet = workbook[record._api]
            counts[record._api] += 1

        else:
            if record._api not in workbook.sheetnames:
                workbook.create_sheet(title=record._api)

            sheet = workbook[record._api]
            sheet.cell(1, 1, 'Message')
            sheet.cell(2, 1, 'Message').alignment = ALIGNMENT_ROTATE
            sheet.cell(3, 1, 'no')

            for index, (key, value) in enumerate(iterable=record.schema()['properties'].items(), start=2):
                print(key, value, '\n\n\n')
                sheet.cell(1, index, key)
                sheet.cell(2, index, value['title']).alignment = ALIGNMENT_ROTATE
                sheet.cell(3, index, 'yes')

            counts[record._api] = 4

        for index, value in enumerate(iterable=record.dict().values(), start=2):
            sheet.cell(counts[record._api], index, value)


    workbook.save(filename=file_path)